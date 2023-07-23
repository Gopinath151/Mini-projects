import mysql.connector   # importing mysql connector
import openpyxl  #imported Openpyxl to export data to excel
from openpyxl import styles

# making Connection
My_DB = mysql.connector.connect(
    host="127.0.0.1", user="root", password="Egopi@151", database="employee")

workbook = openpyxl.Workbook()
worksheet = workbook.active



# sqlquery = "create table emp_management(emp_ID int not null, emp_name char(220), " \
#            "emp_designation char(220), emp_salary bigint)"
# cur = My_DB.cursor()
# cur.execute(sqlquery)

print("Welcome to Employee Management Record\n")


# selecting from user input to perform what action to be done
def Main_menu():

    print("Select the Number")
    print("Select 1 to export Employee")
    print("Select 2 to Add Employee")
    print("Select 3 to Remove Employee ")
    print("Select 4 to Promote Employee")
    print("Select 5 to Exit")


    choice = int(input("Enter your Choice: "))
    if choice == 1:
        View_Employees()
    elif choice == 2:
         Add_Employee()
    elif choice == 3:
        Remove_Employee()
    elif choice == 4:
        Promote_Employee()
    elif choice == 5:
        exit(0)
    else:
        print("Please enter a valid input")
        Main_menu()

def Check_Employee(employee_id):

    sql = 'select * from emp_management where emp_ID=%s'  #sql query
    c = My_DB.cursor(buffered=True)
    data = employee_id
    c.execute(sql, data)

    row = c.rowcount
    if row == 1:
        return True
    else:
        return False


def Add_Employee():
    Id = input("Enter Employee Id : ")

    if (Check_Employee(Id) == True):  #checking if the employee is already present or not
        print("Employee already exists!!!")
        Main_menu()

    else:
        Name = input("Enter Employee Name : ")
        Designation = input("Enter Employee Designation : ")
        Salary = input("Enter Employee Salary : ")
        data = (Id, Name, Designation, Salary)
        sql = 'insert into emp_management values(%s,%s,%s,%s)'
        cursor = My_DB.cursor()

        cursor.execute(sql, data)

        My_DB.commit()
        print("Employee Added Successfully!!\n ")
        Main_menu()

def Remove_Employee():
    Id = input("Enter Employee Id : ")

    if (Check_Employee(Id) == False):  #checking if the employee is already present or not
        print("Employee does not  exists!!!")
        Main_menu()
    else:
        sql = 'delete from emp_management where emp_ID=%s'
        data = (Id,)
        cursor = My_DB.cursor()
        cursor.execute(sql, data)
        My_DB.commit()
        print("Employee Removed!!")
        Main_menu()

# Function to Promote Employee
def Promote_Employee():
    Id = int(input("Enter Employ's Id"))

    if (Check_Employee(Id) == False):  #checking if the employee is already present or not
        print("Employee does not  exists!!!")
        Main_menu()
    else:
        Amount = int(input("Enter increase in Salary"))

        sql = 'select emp_salary from emp_management where emp_ID=%s'
        data = (Id,)
        cursor = My_DB.cursor()
        cursor.execute(sql, data)
        previous_salary = cursor.fetchone()
        print(previous_salary)
        total_salary = previous_salary[0] + Amount

        sql = 'update emp_management set emp_salary=%s where emp_ID=%s'
        update_data = (total_salary, Id)

        cursor.execute(sql, update_data)

        My_DB.commit()
        print("Employee salary updated")
        Main_menu()

def View_Employees():
    sql = 'select * from emp_management'
    cursor = My_DB.cursor()

    cursor.execute(sql)

    data = cursor.fetchall()
    headers = ["ID", "Name","Designation", "salary"]
    for col, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    for row, record in enumerate(data, start=2):
        for col, value in enumerate(record, start=1):
            cell = worksheet.cell(row=row, column=col)
            cell.value = value
            cell.alignment = openpyxl.styles.Alignment(horizontal="center") #implenting styles to export in excel

    print("Employee details is Exported in Excel file")

    workbook.save('C:/Users/Gopinath/OneDrive/Documents/example1.xlsx')  #path to export the excel file
    Main_menu()

Main_menu()

